# src/aje_libs/bd/helpers/excel_helper.py
from openpyxl import load_workbook
from typing import List, Dict, Optional, Any
from ...common.logger import custom_logger

logger = custom_logger(__name__)

class ExcelHelper:
    """Helper para procesar archivos Excel"""
    
    def __init__(self):
        pass
    
    def extract_all_data(self, file_path: str) -> Dict[str, List[List[Any]]]:
        """
        Extrae todos los datos de un archivo Excel.
        
        :param file_path: Ruta al archivo Excel
        :return: Diccionario con datos de todas las hojas
        """
        try:
            workbook = load_workbook(file_path)
            all_data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        sheet_data.append(list(row))
                
                all_data[sheet_name] = sheet_data
            
            logger.info(f"Datos extraídos de {len(all_data)} hojas")
            return all_data
        except Exception as e:
            logger.error(f"Error extrayendo datos de Excel: {e}")
            raise e
    
    def extract_sheet(self, file_path: str, sheet_name: str) -> List[List[Any]]:
        """
        Extrae datos de una hoja específica.
        
        :param file_path: Ruta al archivo Excel
        :param sheet_name: Nombre de la hoja
        :return: Datos de la hoja
        """
        try:
            workbook = load_workbook(file_path)
            sheet = workbook[sheet_name]
            sheet_data = []
            
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    sheet_data.append(list(row))
            
            logger.info(f"Datos extraídos de la hoja '{sheet_name}'")
            return sheet_data
        except KeyError:
            logger.error(f"Hoja '{sheet_name}' no encontrada")
            raise KeyError(f"La hoja '{sheet_name}' no existe")
        except Exception as e:
            logger.error(f"Error extrayendo hoja: {e}")
            raise e
    
    def get_sheet_names(self, file_path: str) -> List[str]:
        """
        Obtiene los nombres de todas las hojas.
        
        :param file_path: Ruta al archivo Excel
        :return: Lista de nombres de hojas
        """
        try:
            workbook = load_workbook(file_path)
            sheet_names = workbook.sheetnames
            logger.info(f"Encontradas {len(sheet_names)} hojas")
            return sheet_names
        except Exception as e:
            logger.error(f"Error obteniendo nombres de hojas: {e}")
            raise e
    
    def extract_with_headers(self, file_path: str, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Extrae datos con encabezados de columna.
        
        :param file_path: Ruta al archivo Excel
        :param sheet_name: Nombre de la hoja
        :return: Lista de diccionarios con datos
        """
        try:
            workbook = load_workbook(file_path)
            sheet = workbook[sheet_name]
            
            # Obtener headers de la primera fila
            headers = []
            for cell in sheet[1]:
                headers.append(cell.value)
            
            # Extraer datos
            data_with_headers = []
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if any(cell is not None for cell in row):
                    row_data = {}
                    for header, value in zip(headers, row):
                        row_data[header] = value
                    data_with_headers.append(row_data)
            
            logger.info(f"Extraídos {len(data_with_headers)} registros con headers")
            return data_with_headers
        except Exception as e:
            logger.error(f"Error extrayendo datos con headers: {e}")
            raise e